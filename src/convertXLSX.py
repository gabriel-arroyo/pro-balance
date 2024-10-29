from openpyxl import load_workbook
import os
import xlwt


def convert_xlsx_to_xls(xlsx_file, xls_file):
    # Open the .xlsx file using openpyxl
    xlsx_workbook = load_workbook(xlsx_file)
    xlsx_sheet = xlsx_workbook.active  # Get the active sheet

    # Create a new .xls workbook and sheet using xlwt
    xls_workbook = xlwt.Workbook()
    xls_sheet = xls_workbook.add_sheet(xlsx_sheet.title)

    # Copy all rows and columns from .xlsx to the new .xls file
    for row in xlsx_sheet.iter_rows():
        for cell in row:
            # Adjust for 0-based indexing
            xls_sheet.write(cell.row - 1, cell.column - 1, cell.value)

    # Save the new .xls file
    xls_workbook.save(xls_file)
    print(f"Conversion successful: {xlsx_file} -> {xls_file}")


# Get the current script directory
current_directory = os.path.dirname(os.path.abspath(__file__))

# Move one level up
parent_directory = os.path.abspath(os.path.join(current_directory, '..'))

# Path to the file one level up
xls_file = os.path.join(parent_directory, 'forms',
                        'template_formato_test1.xlsx')
xlsx_file = os.path.join(parent_directory, 'forms',
                         'template_formato_test1.xls')

convert_xlsx_to_xls(xls_file, xlsx_file)
