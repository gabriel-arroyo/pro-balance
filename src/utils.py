from openpyxl.utils import range_boundaries
from datetime import date, datetime
import os


def get_db_path():
    # Get the current script directory
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Move one level up
    parent_directory = os.path.abspath(os.path.join(current_directory, '..'))

    # Path to the database file
    db_path = os.path.join(parent_directory, 'db', 'inventory.db')
    return db_path


def get_form_path(file):
    # Get the current script directory
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Move one level up
    parent_directory = os.path.abspath(os.path.join(current_directory, '..'))

    # Path to the database file
    form_path = os.path.join(parent_directory, 'forms', file)
    return form_path


def get_report_path(file):
    # Get the current script directory
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Move one level up
    parent_directory = os.path.abspath(os.path.join(current_directory, '..'))

    # Path to the database file
    form_path = os.path.join(parent_directory, 'reports', file)
    return form_path


def get_start_and_end_dates(month_number, year):
    # Get the first day of the month
    start_date = date(year, month_number, 1)

    # Get the first day of the next month
    if month_number == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month_number + 1, 1)

    # Convert dates to strings in 'YYYY-MM-DD' format
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    return start_date_str, end_date_str


def can_write_cell(sheet, row, col):
    # Check if the cell is part of a merged range
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(
            str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            # Return True if it is the top-left cell
            return min_row == row and min_col == col
    return True  # If not merged, it's safe to write


def letter_to_number(letter):
    letter = letter.upper()
    if letter.isalpha() and len(letter) == 1:
        return ord(letter) - ord('A') + 1
    else:
        return "Invalid input. Please enter a single letter."


def number_to_letter(number):
    if 1 <= number <= 26:
        # ASCII value of 'A' is 65, so we add 64 to the number
        return chr(number + 64)
    else:
        return "Invalid input. Please enter a number between 1 and 26."


def get_month_name_spanish(month_number):
    months = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    return months.get(month_number, "Invalid month number")


def insert_cell(sheet, letter, row, value):
    start_letter = letter
    start_row = row
    start_col = letter_to_number(start_letter)
    if isinstance(value, str):
        sheet.cell(start_row, start_col, value.upper())
    else:
        sheet.cell(start_row, start_col, value)


def get_file_names(form_number, month, year):
    month_name = get_month_name_spanish(month)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file_name = f"FORMATO {form_number} SEDENA {
        month_name.upper()} {year}_{timestamp}.xlsx"
    input_file_name = f"template_form_{form_number}.xlsx"
    return get_form_path(input_file_name), get_report_path(output_file_name)


def write_table(sheet, data, start_letter, start_row, has_id=False, transposed=False, formatted_zero=False):
    if not data:
        print("No data to insert on form 1 table 1.")
        return

    start_col = letter_to_number(start_letter)

    if (transposed):
        p = 0
        col = start_col

        while p < len(data):
            if can_write_cell(sheet, start_row, col):
                for i in range(len(data[1])-int(has_id)):
                    value = data[p][i + int(has_id)]
                    formatted_value = value if isinstance(
                        value, str) or value > 0 else 'XXX'
                    sheet.cell(start_row + i, col,
                               value=formatted_value if formatted_zero else value)
                col += 1
                p += 1
            else:
                col += 1
    else:
        for i in range(len(data)):
            j = 0
            col = start_col
            while j < len(data[i])-int(has_id):
                if can_write_cell(sheet, start_row + i, col):
                    value = data[i][j + int(has_id)]
                    formatted_value = value if isinstance(
                        value, str) or value > 0 else 'XXX'
                    sheet.cell(
                        start_row + i, col, value=formatted_value if formatted_zero else value)
                    col += 1
                    j += 1
                else:
                    col += 1


def add_results(sheet, start_letter, start_row, num_rows, num_cols, transposed=False, result_formula=None):
    """
    Adds a results column or row to an existing table in the sheet.

    :param sheet: The worksheet to modify
    :param start_letter: The starting column letter of the existing table
    :param start_row: The starting row number of the existing table
    :param num_rows: Number of rows in the existing table (excluding header)
    :param num_cols: Number of columns in the existing table
    :param transposed: Whether the table is transposed (True for column-wise, False for row-wise)
    :param result_formula: Formula to calculate the result (if None, no formula is added)
    :return: The column letter or row number where results were added
    """
    start_col = letter_to_number(start_letter)

    if transposed:
        # Add results row
        result_row = start_row + num_rows
        p = 0
        col = start_col
        while p < num_cols:
            if can_write_cell(sheet, result_row, col):
                if result_formula and col > start_col:
                    formula = result_formula.format(
                        col_letter=number_to_letter(col),
                        start_row=start_row,
                        end_row=result_row - 1
                    )
                    sheet.cell(result_row, col, formula)
                else:
                    sheet.cell(result_row, col, f"=SUM({number_to_letter(col)}{
                               start_row}:{number_to_letter(col)}{start_row + num_rows - 1})")
                p += 1
                col += 1
            else:
                col += 1

        return result_row
    else:
        # Add results column
        result_col = start_col + num_cols
        result_col_letter = number_to_letter(result_col)

        for row in range(start_row, start_row + num_rows + 1):
            if can_write_cell(sheet, row, result_col):
                if result_formula and row > start_row:
                    formula = result_formula.format(
                        row_num=row,
                        start_col=number_to_letter(start_col),
                        end_col=number_to_letter(result_col - 1)
                    )
                    sheet.cell(row, result_col, formula)
                else:
                    sheet.cell(row, result_col, "")

        return result_col_letter

# Example usage:
# result_formula = "=AVERAGE({start_col}{row_num}:{end_col}{row_num})"
# add_results(sheet, 'A', 1, 10, 5, transposed=False, result_formula=result_formula)



def format_date_spanish(date):
    # Spanish months in uppercase
    months_spanish = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    
    day = date.day
    month = months_spanish[date.month - 1]  # Get the month in Spanish
    year = date.year
    
    return f"{day} DE {month} DEL {year}"