import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def generate_excel_report():
    # Connect to the SQLite database
    conn = sqlite3.connect('inventory.db')
    cursor = conn.cursor()

    # Query to get the data for the report
    query = """
    SELECT p.name AS Product, t.type AS TransactionType, t.quantity, t.date
    FROM Transactions t
    JOIN Products p ON t.product_id = p.id
    """
    cursor.execute(query)
    data = cursor.fetchall()

    # Create a new Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Report"

    # Add headers
    headers = ['Product', 'Transaction Type', 'Quantity', 'Date']
    sheet.append(headers)

    # Apply formatting to the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD",
                              end_color="4F81BD", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add data to the sheet
    for row in data:
        sheet.append(row)

    # Auto-adjust column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # **Save the workbook to a specified location**
    report_filename = "monthly_inventory_report.xlsx"
    workbook.save(report_filename)
    print(f"Report saved successfully as '{report_filename}'")


generate_excel_report()
